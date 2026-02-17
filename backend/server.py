from fastapi import FastAPI, APIRouter, Depends, HTTPException, status, Header, Request
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import io

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ.get('JWT_SECRET', 'gatekeeper-secret-key-2024')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

app = FastAPI()
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ─── Models ───────────────────────────────────────────────────────────

class UserCreate(BaseModel):
    username: str
    password: str
    name: str
    role: str = "porteiro"

class UserUpdate(BaseModel):
    username: Optional[str] = None
    password: Optional[str] = None
    name: Optional[str] = None
    role: Optional[str] = None

class LoginRequest(BaseModel):
    username: str
    password: str

class VisitorCreate(BaseModel):
    name: str
    document: str
    entry_time: Optional[str] = None
    vehicle_plate: Optional[str] = ""
    company: Optional[str] = ""
    observation: Optional[str] = ""

class ScheduleCreate(BaseModel):
    visitor_name: str
    company: Optional[str] = ""
    visit_date: str
    visit_time: str
    notes: Optional[str] = ""

class FleetTripCreate(BaseModel):
    driver_name: str
    vehicle: str
    departure_km: float

class FleetTripReturn(BaseModel):
    arrival_km: float

class ReportObservation(BaseModel):
    observation: str
    porter_name: str

# ─── Auth Helpers ─────────────────────────────────────────────────────

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, username: str, role: str, name: str) -> str:
    payload = {
        "user_id": user_id,
        "username": username,
        "role": role,
        "name": name,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request):
    token = request.headers.get("authorization")
    if not token:
        token = request.query_params.get("authorization")
    if not token:
        raise HTTPException(status_code=401, detail="Token não fornecido")
    try:
        if token.startswith("Bearer "):
            token = token[7:]
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")

# ─── Startup ──────────────────────────────────────────────────────────

@app.on_event("startup")
async def startup():
    admin = await db.users.find_one({"username": "admin"}, {"_id": 0})
    if not admin:
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "username": "admin",
            "password": hash_password("admin123"),
            "name": "Administrador",
            "role": "admin",
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info("Admin padrão criado: admin / admin123")
    # Create indexes
    await db.visitors.create_index("entry_time")
    await db.schedules.create_index("visit_date")
    await db.fleet_trips.create_index("created_at")

# ─── Auth Routes ──────────────────────────────────────────────────────

@api_router.post("/auth/login")
async def login(req: LoginRequest):
    user = await db.users.find_one({"username": req.username}, {"_id": 0})
    if not user or not verify_password(req.password, user["password"]):
        raise HTTPException(status_code=401, detail="Credenciais inválidas")
    token = create_token(user["id"], user["username"], user["role"], user["name"])
    return {"token": token, "user": {"id": user["id"], "username": user["username"], "name": user["name"], "role": user["role"]}}

@api_router.get("/auth/verify")
async def verify_token(request: Request):
    user = await get_current_user(request)
    return {"user": {"id": user["user_id"], "username": user["username"], "name": user["name"], "role": user["role"]}}

# ─── User Management (Admin only) ────────────────────────────────────

@api_router.get("/users")
async def list_users(request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Acesso negado")
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    return users

@api_router.post("/users")
async def create_user(req: UserCreate, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Acesso negado")
    existing = await db.users.find_one({"username": req.username}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Usuário já existe")
    new_user = {
        "id": str(uuid.uuid4()),
        "username": req.username,
        "password": hash_password(req.password),
        "name": req.name,
        "role": req.role,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(new_user)
    return {"id": new_user["id"], "username": new_user["username"], "name": new_user["name"], "role": new_user["role"]}

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, req: UserUpdate, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Acesso negado")
    update_data = {}
    if req.username:
        update_data["username"] = req.username
    if req.password:
        update_data["password"] = hash_password(req.password)
    if req.name:
        update_data["name"] = req.name
    if req.role:
        update_data["role"] = req.role
    if not update_data:
        raise HTTPException(status_code=400, detail="Nenhum campo para atualizar")
    result = await db.users.update_one({"id": user_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    return {"message": "Usuário atualizado"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, request: Request):
    user = await get_current_user(request)
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Acesso negado")
    if user["user_id"] == user_id:
        raise HTTPException(status_code=400, detail="Não pode deletar a si mesmo")
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    return {"message": "Usuário deletado"}

# ─── Visitors ─────────────────────────────────────────────────────────

@api_router.post("/visitors")
async def create_visitor(req: VisitorCreate, request: Request):
    await get_current_user(request)
    visitor = {
        "id": str(uuid.uuid4()),
        "name": req.name,
        "document": req.document,
        "entry_time": req.entry_time or datetime.now(timezone.utc).isoformat(),
        "exit_time": None,
        "vehicle_plate": req.vehicle_plate or "",
        "company": req.company or "",
        "observation": req.observation or "",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.visitors.insert_one(visitor)
    return {k: v for k, v in visitor.items() if k != "_id"}

@api_router.get("/visitors")
async def list_visitors(request: Request, date: Optional[str] = None, active: Optional[bool] = None):
    await get_current_user(request)
    query = {}
    if active is True:
        query["exit_time"] = None
    if date:
        query["entry_time"] = {"$regex": f"^{date}"}
    visitors = await db.visitors.find(query, {"_id": 0}).sort("entry_time", -1).to_list(1000)
    return visitors

@api_router.put("/visitors/{visitor_id}/checkout")
async def checkout_visitor(visitor_id: str, request: Request):
    await get_current_user(request)
    exit_time = datetime.now(timezone.utc).isoformat()
    result = await db.visitors.update_one(
        {"id": visitor_id, "exit_time": None},
        {"$set": {"exit_time": exit_time}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Visitante não encontrado ou já deu saída")
    return {"message": "Saída registrada", "exit_time": exit_time}

# ─── Schedules ────────────────────────────────────────────────────────

@api_router.post("/schedules")
async def create_schedule(req: ScheduleCreate, request: Request):
    await get_current_user(request)
    schedule = {
        "id": str(uuid.uuid4()),
        "visitor_name": req.visitor_name,
        "company": req.company or "",
        "visit_date": req.visit_date,
        "visit_time": req.visit_time,
        "notes": req.notes or "",
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.schedules.insert_one(schedule)
    return {k: v for k, v in schedule.items() if k != "_id"}

@api_router.get("/schedules")
async def list_schedules(request: Request, date: Optional[str] = None):
    await get_current_user(request)
    query = {}
    if date:
        query["visit_date"] = date
    schedules = await db.schedules.find(query, {"_id": 0}).sort("visit_date", 1).to_list(1000)
    return schedules

@api_router.get("/schedules/today")
async def get_today_schedules(request: Request):
    await get_current_user(request)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    schedules = await db.schedules.find({"visit_date": today, "status": "pending"}, {"_id": 0}).to_list(1000)
    return schedules

@api_router.put("/schedules/{schedule_id}/complete")
async def complete_schedule(schedule_id: str, request: Request):
    await get_current_user(request)
    result = await db.schedules.update_one({"id": schedule_id}, {"$set": {"status": "completed"}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Agendamento não encontrado")
    return {"message": "Agendamento concluído"}

@api_router.delete("/schedules/{schedule_id}")
async def delete_schedule(schedule_id: str, request: Request):
    await get_current_user(request)
    result = await db.schedules.delete_one({"id": schedule_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Agendamento não encontrado")
    return {"message": "Agendamento deletado"}

# ─── Fleet ────────────────────────────────────────────────────────────

@api_router.post("/fleet")
async def create_fleet_trip(req: FleetTripCreate, request: Request):
    await get_current_user(request)
    trip = {
        "id": str(uuid.uuid4()),
        "driver_name": req.driver_name,
        "vehicle": req.vehicle,
        "departure_km": req.departure_km,
        "arrival_km": None,
        "distance": None,
        "status": "em_viagem",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.fleet_trips.insert_one(trip)
    return {k: v for k, v in trip.items() if k != "_id"}

@api_router.get("/fleet")
async def list_fleet_trips(request: Request, date: Optional[str] = None, active: Optional[bool] = None):
    await get_current_user(request)
    query = {}
    if active is True:
        query["status"] = "em_viagem"
    if date:
        query["created_at"] = {"$regex": f"^{date}"}
    trips = await db.fleet_trips.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return trips

@api_router.put("/fleet/{trip_id}/return")
async def return_fleet_trip(trip_id: str, req: FleetTripReturn, request: Request):
    await get_current_user(request)
    trip = await db.fleet_trips.find_one({"id": trip_id}, {"_id": 0})
    if not trip:
        raise HTTPException(status_code=404, detail="Viagem não encontrada")
    if trip["status"] != "em_viagem":
        raise HTTPException(status_code=400, detail="Veículo já retornou")
    distance = req.arrival_km - trip["departure_km"]
    result = await db.fleet_trips.update_one(
        {"id": trip_id},
        {"$set": {"arrival_km": req.arrival_km, "distance": distance, "status": "retornado"}}
    )
    return {"message": "Retorno registrado", "distance": distance}

# ─── Reports ──────────────────────────────────────────────────────────

@api_router.get("/reports/daily")
async def get_daily_report(request: Request, date: Optional[str] = None):
    await get_current_user(request)
    if not date:
        date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    visitors = await db.visitors.find({"entry_time": {"$regex": f"^{date}"}}, {"_id": 0}).to_list(1000)
    fleet = await db.fleet_trips.find({"created_at": {"$regex": f"^{date}"}}, {"_id": 0}).to_list(1000)
    schedules = await db.schedules.find({"visit_date": date}, {"_id": 0}).to_list(1000)
    report_obs = await db.report_observations.find_one({"date": date}, {"_id": 0})
    return {
        "date": date,
        "visitors": visitors,
        "fleet": fleet,
        "schedules": schedules,
        "observation": report_obs.get("observation", "") if report_obs else "",
        "porter_name": report_obs.get("porter_name", "") if report_obs else ""
    }

@api_router.post("/reports/observation")
async def save_report_observation(req: ReportObservation, request: Request, date: Optional[str] = None):
    await get_current_user(request)
    if not date:
        date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    await db.report_observations.update_one(
        {"date": date},
        {"$set": {"date": date, "observation": req.observation, "porter_name": req.porter_name, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    return {"message": "Observação salva"}

@api_router.get("/reports/export/excel")
async def export_excel(request: Request, date: Optional[str] = None):
    await get_current_user(request)
    if not date:
        date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório Diário"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    sub_header_font = Font(bold=True, size=11)
    sub_header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    row = 1
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = f"RELATÓRIO DIÁRIO - PORTARIA - {date}"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 2

    # Visitors
    visitors = await db.visitors.find({"entry_time": {"$regex": f"^{date}"}}, {"_id": 0}).to_list(1000)
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "VISITANTES"
    ws[f'A{row}'].font = sub_header_font
    ws[f'A{row}'].fill = sub_header_fill
    row += 1
    headers = ["Nome", "Documento", "Entrada", "Saída", "Placa", "Empresa", "Observação"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row += 1
    for v in visitors:
        entry = v.get("entry_time", "")[:16].replace("T", " ") if v.get("entry_time") else ""
        exit_t = v.get("exit_time", "")[:16].replace("T", " ") if v.get("exit_time") else "Em andamento"
        vals = [v.get("name",""), v.get("document",""), entry, exit_t, v.get("vehicle_plate",""), v.get("company",""), v.get("observation","")]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
        row += 1
    row += 1

    # Fleet
    fleet = await db.fleet_trips.find({"created_at": {"$regex": f"^{date}"}}, {"_id": 0}).to_list(1000)
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = "CONTROLE DE FROTA"
    ws[f'A{row}'].font = sub_header_font
    ws[f'A{row}'].fill = sub_header_fill
    row += 1
    fleet_headers = ["Motorista", "Veículo", "KM Saída", "KM Entrada", "Distância (KM)", "Status"]
    for col, h in enumerate(fleet_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    row += 1
    for f in fleet:
        status_text = "Retornado" if f.get("status") == "retornado" else "Em viagem"
        vals = [f.get("driver_name",""), f.get("vehicle",""), f.get("departure_km",0), f.get("arrival_km","—"), f.get("distance","—"), status_text]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
        row += 1
    row += 1

    # Observations
    report_obs = await db.report_observations.find_one({"date": date}, {"_id": 0})
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "OBSERVAÇÕES DO DIA"
    ws[f'A{row}'].font = sub_header_font
    ws[f'A{row}'].fill = sub_header_fill
    row += 1
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = report_obs.get("observation", "Nenhuma observação") if report_obs else "Nenhuma observação"
    row += 2
    ws[f'A{row}'] = "Porteiro responsável:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = report_obs.get("porter_name", "—") if report_obs else "—"

    # Adjust column widths
    for col in range(1, 8):
        ws.column_dimensions[chr(64+col)].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=relatorio_{date}.xlsx"}
    )

@api_router.get("/reports/export/pdf")
async def export_pdf(request: Request, date: Optional[str] = None):
    await get_current_user(request)
    if not date:
        date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], fontSize=16, spaceAfter=12)
    elements.append(Paragraph(f"RELATÓRIO DIÁRIO - PORTARIA - {date}", title_style))
    elements.append(Spacer(1, 12))

    # Visitors
    visitors = await db.visitors.find({"entry_time": {"$regex": f"^{date}"}}, {"_id": 0}).to_list(1000)
    elements.append(Paragraph("VISITANTES", styles['Heading2']))
    v_data = [["Nome", "Documento", "Entrada", "Saída", "Placa", "Empresa", "Obs."]]
    for v in visitors:
        entry = v.get("entry_time", "")[:16].replace("T", " ") if v.get("entry_time") else ""
        exit_t = v.get("exit_time", "")[:16].replace("T", " ") if v.get("exit_time") else "Em andamento"
        v_data.append([v.get("name",""), v.get("document",""), entry, exit_t, v.get("vehicle_plate",""), v.get("company",""), v.get("observation","")[:30]])
    if len(v_data) == 1:
        v_data.append(["Nenhum visitante registrado", "", "", "", "", "", ""])
    t = RLTable(v_data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 18))

    # Fleet
    fleet = await db.fleet_trips.find({"created_at": {"$regex": f"^{date}"}}, {"_id": 0}).to_list(1000)
    elements.append(Paragraph("CONTROLE DE FROTA", styles['Heading2']))
    f_data = [["Motorista", "Veículo", "KM Saída", "KM Entrada", "Distância (KM)", "Status"]]
    for f in fleet:
        status_text = "Retornado" if f.get("status") == "retornado" else "Em viagem"
        f_data.append([f.get("driver_name",""), f.get("vehicle",""), str(f.get("departure_km",0)), str(f.get("arrival_km","—")), str(f.get("distance","—")), status_text])
    if len(f_data) == 1:
        f_data.append(["Nenhum registro", "", "", "", "", ""])
    t2 = RLTable(f_data, repeatRows=1)
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
    ]))
    elements.append(t2)
    elements.append(Spacer(1, 18))

    # Observations
    report_obs = await db.report_observations.find_one({"date": date}, {"_id": 0})
    elements.append(Paragraph("OBSERVAÇÕES DO DIA", styles['Heading2']))
    obs_text = report_obs.get("observation", "Nenhuma observação") if report_obs else "Nenhuma observação"
    elements.append(Paragraph(obs_text, styles['Normal']))
    elements.append(Spacer(1, 12))
    porter = report_obs.get("porter_name", "—") if report_obs else "—"
    elements.append(Paragraph(f"<b>Porteiro responsável:</b> {porter}", styles['Normal']))

    doc.build(elements)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=relatorio_{date}.pdf"}
    )

# ─── Dashboard Stats ─────────────────────────────────────────────────

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(request: Request):
    await get_current_user(request)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    active_visitors = await db.visitors.count_documents({"exit_time": None})
    today_visitors = await db.visitors.count_documents({"entry_time": {"$regex": f"^{today}"}})
    today_schedules = await db.schedules.count_documents({"visit_date": today, "status": "pending"})
    active_trips = await db.fleet_trips.count_documents({"status": "em_viagem"})
    today_trips = await db.fleet_trips.count_documents({"created_at": {"$regex": f"^{today}"}})
    return {
        "active_visitors": active_visitors,
        "today_visitors": today_visitors,
        "today_schedules": today_schedules,
        "active_trips": active_trips,
        "today_trips": today_trips
    }

# ─── Root ─────────────────────────────────────────────────────────────

@api_router.get("/")
async def root():
    return {"message": "Gatekeeper API Running"}

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
